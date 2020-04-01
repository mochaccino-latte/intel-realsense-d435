#!/usr/bin/env python
import rospy
from std_msgs.msg import String
from std_msgs.msg import UInt16
from sensor_msgs.msg import JointState
import time
import getch
import numpy as np
from openpyxl import workbook
from openpyxl import load_workbook

class ur5_sensor:
    def __init__(self):
        self.something = 0
        self.filepath = "/home/s/catkin_ws/src/camera/src/joints.xlsx"
        self.wb = load_workbook(self.filepath)
        self.sheet = self.wb.active
        self.excel_row = 1
        self.index = 1
        self.save_status = False
        self.joint_states_pos = 0
        self.joint_states_vel = 0
        self.joint_states_eff = 0
        self.command = False
        self.time_init = 0
        self.time_record = 0

    def callback(self,data):
        if self.command == True:
            rospy.loginfo('rec')
            # self.joint_states_pos = [round(i,4) for i in self.joint_states_pos]
            # self.joint_states_vel = [round(i,4) for i in self.joint_states_vel]
            # self.joint_states_eff = [round(i,4) for i in self.joint_states_eff]
            # rospy.loginfo('command: %s', self.joint_states_pos)
            # rospy.loginfo('command: %s', self.joint_states_vel)
            # rospy.loginfo('command: %s', self.joint_states_eff)
            self.time_record = time.time() - self.time_init
            self.sheet.cell(row=self.excel_row, column = 1, value = str(data.position))
            self.sheet.cell(row=self.excel_row, column = 2, value = str(data.velocity))
            self.sheet.cell(row=self.excel_row, column = 3, value = str(data.effort))
            self.sheet.cell(row=self.excel_row, column = 4, value = str(self.time_record))
            self.excel_row+=1

    def callback_command(self,data):
        if (data.data == 114):
            self.time_init = time.time()
            self.command = True
            self.save_status = False
            self.excel_row = 1
        elif (data.data == 115):
            if self.save_status == False :
                rospy.loginfo('saved')
                save_filepath = "/home/s/catkin_ws/src/camera/src/joints/joints" + str(self.index) +  ".xlsx"
                self.wb.save(save_filepath)
                self.index+=1
                self.wb = load_workbook(self.filepath)
                self.sheet = self.wb.active
                self.command = False
                self.save_status = True

    def jointState(self):
        rospy.init_node('listener_to_joint_states', anonymous=True)
        rate = rospy.Rate(30)
        rospy.Subscriber('joint_states', JointState, self.callback)
        rospy.Subscriber('command', UInt16, self.callback_command)
        rospy.spin()


ur5_sensor = ur5_sensor()
if __name__ == '__main__':
    ur5_sensor.jointState()
