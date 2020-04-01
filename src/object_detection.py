#!/usr/bin/env python2
import sys
# sys.path.remove('/opt/ros/kinetic/lib/python2.7/dist-packages')
import cv2
import time
import math
import numpy as np
import matplotlib.pyplot as plt
import pyrealsense2 as rs
from openpyxl import Workbook
# from universal_robot_kinematics import invKine
# from forward_kinematics import fwd_kin
from scipy.spatial.transform import Rotation as R
from pyquaternion import Quaternion
print("Environment Ready")

class IntelRealsense:
    def __init__(self):
        # initiate the pipeline
        self.shift_x = 300
        self.shift_y = 0
        self.pp_k = 0
        self.name = 1
        self.colorizer = rs.colorizer()
        self.pipe = rs.pipeline()
        config = rs.config()
        config.enable_stream(rs.stream.depth, 848, 480, rs.format.z16, 90)
        config.enable_stream(rs.stream.color, 848, 480, rs.format.bgr8, 30)
        profile = self.pipe.start(config)
        i = profile.get_stream(rs.stream.depth)
        self.intr = i.as_video_stream_profile().get_intrinsics()
        # extr = i.as_video_stream_profile().get_extrinsics_to()
        # print('intr %s' %self.intr)
        s = profile.get_device().query_sensors()[1]
        s.set_option(rs.option.enable_auto_exposure, False)
        self.depth_scale = profile.get_device().first_depth_sensor().get_depth_scale()

    def reverse_perspective_projection(self, pp_x, pp_y, pp_depth):
        intr = self.intr
        image_plane = np.array([  [pp_x * pp_depth],
                                  [pp_y * pp_depth],
                                  [1 * pp_depth]      ])
        # print('Image Plane \n %s \n' %image_plane)
        T_perspective_projection = np.array([   [intr.fx, 0,        intr.ppx,      0],
                                                [0,       intr.fy,  intr.ppy,      0],
                                                [0,       0,        1,             0] ])
        # T_perspective_projection * vec(world_coordinates) = image_plane
        answer = np.linalg.lstsq(T_perspective_projection, image_plane, rcond=None)
        return np.array(answer[0])

    # def transformation_image2camera(self, pp_x, pp_y, pp_depth, pp_k, shift_x=0, shift_y=0):
    #     pp_x -= 424; pp_y -= 240;
    #     print('pp camera frame (%s, %s)' %(pp_x, pp_y))
    #     pingpong_camera = np.array([  [pp_x * pp_k *100],
    #                                   [pp_y * pp_k *100],
    #                                   [pp_depth * 100],
    #                                   [1]  ])
    #     return pingpong_camera

    def pingpong_detection(self, shift_x=0, shiftend_x=150, shift_y=75, scope_side=45, scope_depth=50, display=True):
        pp_area = 0.001256 # meters
        capture = time.time()
        frameset = self.pipe.wait_for_frames()
        depth_frame = frameset.get_depth_frame()
        depth = np.asanyarray(depth_frame.get_data())
        depth_shape = depth.shape
        depth_crop = depth[0:depth_shape[0]-shift_y, 0+shift_x:depth_shape[1]-shiftend_x]
        min = depth_crop[depth_crop > 10].min()     # Depth Values
        # print('depth : %f' %(min))
        if min > 700 and min < 2500:
            min_pt = np.where(depth_crop == min)
            depth_scope = depth_crop[int(min_pt[0][0]-scope_side/2):int(min_pt[0][0]+scope_side/2), int(min_pt[1][0]-scope_side/2): int(min_pt[1][0]+scope_side/2)]
            numpix = 0; sumx = 0; sumy = 0
            for row in range(0,depth_scope.shape[0]):
                for col in range(0,depth_scope.shape[1]):
                    if depth_scope[row,col] < min+scope_depth and depth_scope[row,col] > min-scope_depth:numpix+=1; sumx += col; sumy += row
            if numpix != 0:ppscope_x = sumx/numpix; ppscope_y = sumy/numpix
            else:ppscope_x = 0; ppscope_y = 0
            pp_x = ppscope_x+shift_x+min_pt[1][0]-scope_side/2
            pp_y = ppscope_y+min_pt[0][0]-scope_side/2
            pp_depth = depth[int(pp_y),int(pp_x)] * self.depth_scale
            if numpix != 0:
                self.pp_k = math.sqrt(pp_area/numpix)
        else:
            pp_x=0; pp_y=0; pp_depth=0
        if display == True:
            depth_color = np.asanyarray(self.colorizer.colorize(depth_frame).get_data())
            cv2.line(depth_color, (shift_x,0), (shift_x, depth_shape[0]), (0,0,255), 1)
            cv2.line(depth_color, (0,depth_shape[0]-shift_y), (depth_shape[1], depth_shape[0]-shift_y), (0,0,255), 1)
            cv2.circle(depth_color, (int(pp_x),int(pp_y)), (int)(10),(0,255,0),1)
            cv2.circle(depth_color, (int(np.size(depth,1)/2),int(np.size(depth,0)/2)), (int)(0),(0,0,255),5)
            cv2.namedWindow('Object Detectiom using Depth Image', cv2.WINDOW_AUTOSIZE)
            cv2.imshow('Object Detectiom using Depth Image', depth_color)
            # filename = 'pingpong'+str(self.name)+'.png'
            # cv2.imwrite('/home/idealab/catkin_ws/src/thesis/src/detect/'+filename,depth_color)
            # cv2.waitKey(0)
            key = cv2.waitKey(1)
            if key & 0xFF == ord('q') or key == 27:
                cv2.destroyAllWindows()
        return pp_x, pp_y, pp_depth, self.pp_k, capture

    def pingpong_velocity(self, STATE, pp_x, pp_y, pp_depth, capture, display=True):
        v_x = 999; v_y = 999; v_depth = 999;a_x =999; a_y = 999; a_depth =999
        if pp_x == 0 and pp_y == 0:
            STATE = 'NONE'
        else:
            if STATE == 'NONE':STATE='INITIAL'
            if STATE == 'INITIAL':
                self.lastpp_x = pp_x; self.lastpp_y = pp_y; self.lastcapture = capture
                self.lastpp_depth = pp_depth; STATE = 'VELOCITY'
            elif STATE == 'VELOCITY':
                delt = capture-self.lastcapture
                v_x = (pp_x - self.lastpp_x)/(delt); v_y = (pp_y - self.lastpp_y)/(delt)
                v_depth = (pp_depth - self.lastpp_depth)/(delt)
                self.lastv_x = v_x; self.lastv_y = v_y
                self.lastv_depth = v_depth
                self.lastpp_x = pp_x; self.lastpp_y = pp_y
                self.lastpp_depth = pp_depth
                STATE = 'KALMAN'
            elif STATE == 'KALMAN':
                delt = capture-self.lastcapture
                v_x = (pp_x - self.lastpp_x)/(delt); v_y = (pp_y - self.lastpp_y)/(delt)
                v_depth = (pp_depth - self.lastpp_depth)/(delt)
                a_x = (v_x - self.lastv_x)/delt; a_y = (v_y - self.lastv_y)/delt
                a_depth = (v_depth - self.lastv_depth)/(delt)
                if display != True:
                    self.lastv_x = v_x; self.lastv_y = v_y
                    self.lastv_depth = v_depth
                    self.lastpp_x = pp_x; self.lastpp_y = pp_y
                    self.lastpp_depth = pp_depth
                    self.lasta_x = a_x;self.lasta_y = a_y
                    self.lasta_depth = a_depth
        if display == True and  STATE == 'KALMAN' :
            frameset = self.pipe.wait_for_frames()
            depth_frame = frameset.get_depth_frame()
            depth_color = np.asanyarray(self.colorizer.colorize(depth_frame).get_data())
            predpp_x = self.lastpp_x + self.lastv_x*delt
            # predpp_y = self.lastpp_y + self.lastv_y*delt + 0.5*9.8*(delt**2)
            self.lastv_x = v_x; self.lastv_y = v_y
            self.lastpp_x = pp_x; self.lastpp_y = pp_y
            cv2.line(depth_color, (int(predpp_x), 0), (int(predpp_x), depth_color.shape[0]), (0,255,0), 1)
            # cv2.line(depth_color, (0, int(predpp_y)), (depth_color.shape[1],int(predpp_y)), (0,255,0), 1)
            cv2.namedWindow('State Prediction', cv2.WINDOW_AUTOSIZE)
            cv2.imshow('State Prediction', depth_color)
            filename = 'pingpong'+str(self.name)+'.png'
            cv2.imwrite('/home/idealab/catkin_ws/src/thesis/src/predict/'+filename,depth_color)
            self.name+=1
            key = cv2.waitKey(1)
            if key & 0xFF == ord('q') or key == 27:
                cv2.destroyAllWindows()
        return v_x,v_y,v_depth,a_x,a_y,a_depth,STATE

    def transformation_camera2base(self, x_camera=-1.10, y_camera=1.5, z_camera=0.31):
        T_camera2base = np.array([  [-1, 0, 0, x_camera],
                                    [0, 0, -1, y_camera],
                                    [0, -1, 0, z_camera],
                                    [0, 0, 0, 1]])
        return T_camera2base

    def transformation_end2base(self, x_end=0, y_end=0, z_end=0):
        T_end2base = np.array([  [0, 0, -1, x_end],
                                 [1, 0, 0, y_end],
                                 [0, -1, 0, z_end],
                                 [0, 0, 0, 1]  ])
        return T_end2base

# __main__
obj_x = 10; obj_y = 10; obj_z=10
MAT = [[obj_x],
        [obj_y],
        [obj_z],
        [1] ]
IntelRealsense = IntelRealsense()

if __name__ == '__main__':
    # Prepare for excel
    book = Workbook()
    sheet = book.active
    excel_row = 1
    print("Initiate Object Detection")
    STATE = 'INITIAL'
    _,_,_,_, last_capture = IntelRealsense.pingpong_detection(display = False)
while True:
    # processing time
    initiate_process_time = time.time()
    pp_x, pp_y, pp_depth, pp_k, capture = IntelRealsense.pingpong_detection(display = True)
    processing = capture - last_capture
    # print('proccessing time %f' %(processing))
    v_x, v_y, v_depth, a_x, a_y, a_depth, STATE = IntelRealsense.pingpong_velocity(STATE, pp_x, pp_y, pp_depth, capture, display = False)
    # print('pp_x, pp_y, v_x, v_y : %.3f %.3f %s %s %s %s %s' %(pp_x,pp_y,v_x,v_y,v_depth,a_x,a_y))
    # sheet.cell(row = excel_row,column = 1).value = pp_x
    # sheet.cell(row = excel_row,column = 2).value = pp_y
    # sheet.cell(row = excel_row,column = 3).value = pp_depth
    # sheet.cell(row = excel_row,column = 4).value = v_x
    # sheet.cell(row = excel_row,column = 5).value = v_y
    # sheet.cell(row = excel_row,column = 6).value = v_depth
    # sheet.cell(row = excel_row,column = 7).value = a_x
    # sheet.cell(row = excel_row,column = 8).value = a_y
    # sheet.cell(row = excel_row,column = 9).value = a_depth
    # print('excel row : %s' %excel_row)
    # if excel_row == 10000:
    #     book.save('/home/s/catkin_ws/src/camera/src/variance.xlsx')
    #     print('excel saved')
    last_capture = capture
    final_process_time = time.time()
    process_time = final_process_time-initiate_process_time
    sheet.cell(row = excel_row,column = 1).value = process_time
    excel_row +=1
    print('Process Time : %s'%(final_process_time-initiate_process_time))
    if excel_row == 1000:
        book.save('/home/s/catkin_ws/src/camera/src/process_time.xlsx')
    # position = fwd_kin([math.pi/4, -math.pi/4, math.pi/4, -math.pi/4, math.pi/4, math.pi/4], o_unit='n')
    # print('fwd_kin',position)
    # INV = invKine(np.matrix(position))
    # DEG = (180/math.pi)*INV
    # print('invKine',INV)
    # print('\n')
